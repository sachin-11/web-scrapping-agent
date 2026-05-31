import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyxlImage

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

def create_excel(data: list, image_paths: dict = None, output_path: str = None) -> str:
    """
    Generates a professionally styled Excel spreadsheet from a list of dicts.
    Features frozen headers, autofilter, column auto-width, and optional image embedding.
    """
    if not data:
        logger.warning("No data provided to create_excel.")
        return ""

    # Set default output path
    if not output_path:
        backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        output_path = os.path.join(backend_dir, "outputs", "result.xlsx")

    # Ensure parent output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Initialize Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    # Auto-detect columns (headers) from keys of first dictionary
    headers = list(data[0].keys())
    
    # Check if we need to add an "Image" column
    include_images = image_paths and len(image_paths) > 0
    if include_images:
        headers.append("Image")

    # Header styling configurations
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # Dark Blue (#1E40AF)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Grid lines / Borders
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header).upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze the top header row
    ws.freeze_panes = "A2"

    # Zebra striping backgrounds
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_gray = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Light Gray (#F3F4F6)
    
    row_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")

    # Write data rows
    for row_idx, item in enumerate(data, 2):
        # Alternate backgrounds
        row_fill = fill_gray if row_idx % 2 == 0 else fill_white
        
        # Determine if we should set row height for images
        if include_images:
            ws.row_dimensions[row_idx].height = 60 # 60 points corresponds to 80 pixels (80 / 1.33)
        else:
            ws.row_dimensions[row_idx].height = 20 # Standard height

        # Write product attributes
        for col_idx, header in enumerate(headers[:-1] if include_images else headers, 1):
            val = item.get(header, "")
            # Convert list/dict values to string representation
            if isinstance(val, (list, dict)):
                val = json.dumps(val) if 'json' in globals() else str(val)
                
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = row_font
            cell.fill = row_fill
            cell.alignment = data_alignment
            cell.border = thin_border

        # Embed actual image if provided
        if include_images:
            image_col_idx = len(headers)
            # Alternate background for the image cell
            img_cell = ws.cell(row=row_idx, column=image_col_idx)
            img_cell.fill = row_fill
            img_cell.border = thin_border
            
            # Lookup image by index (both 0-indexed or 1-indexed / strings)
            data_index = row_idx - 2 # 0-indexed
            img_path = (
                image_paths.get(data_index) or 
                image_paths.get(str(data_index)) or 
                image_paths.get(row_idx - 1) or # 1-indexed
                image_paths.get(str(row_idx - 1))
            )
            
            if img_path and os.path.exists(img_path):
                try:
                    logger.info(f"Embedding image: {img_path} at cell coordinate row {row_idx}")
                    openpyxl_img = OpenPyxlImage(img_path)
                    
                    # Force sizing to 80x80 pixels
                    openpyxl_img.width = 80
                    openpyxl_img.height = 80
                    
                    # Target cell coordinate (e.g. "E2")
                    cell_coord = f"{get_column_letter(image_col_idx)}{row_idx}"
                    ws.add_image(openpyxl_img, cell_coord)
                except Exception as e:
                    logger.error(f"Failed to embed image {img_path}: {e}")

    # Set column auto-widths based on contents
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't auto-fit width for the Image column; set it to a fixed size
        if include_images and col[0].column == len(headers):
            ws.column_dimensions[col_letter].width = 15
            continue

        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        # Buffer padding of 3, minimum 12, maximum 50 characters to prevent huge columns
        ws.column_dimensions[col_letter].width = max(12, min(50, max_len + 3))

    # Add filter dropdowns on all columns
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(data) + 1}"

    # Save to path
    logger.info(f"Saving Excel spreadsheet to: {output_path}")
    wb.save(output_path)
    return output_path
